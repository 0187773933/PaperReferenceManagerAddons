from openpyxl import load_workbook
from tqdm import tqdm

from ..utils import utils

def safe_sheet_name( name ):
	safe = name
	for ch in ":/\\?*[]":
		safe = safe.replace( ch , " " )
	return safe.strip()[ :31 ]

class OpenAlexSearch():
	def __init__( self , args ):
		self.args = args
		self.HEADERS_ROW = [ "Cites" , "Title" , "Year" , "Proxy" , "DOI" , "Link" , "OA Cited-By" , "WID" ]
		self.xlsx_path = self.args.output.joinpath( "missing.xlsx" )

	def add_search_sheets( self , oa_index , searches ):
		wb = load_workbook( self.xlsx_path )

		for name , predicate in searches:
			matched = []
			for wid , meta , hay , cite_count , included in tqdm( oa_index , desc=f"Search: {name}" ):
				if not included:
					continue
				if not predicate( hay ):
					continue
				matched.append( utils.openalex_to_xlsx_row( wid , meta , cite_count ) )
			matched.sort( key=lambda r: r[ 2 ] or 0 , reverse=True )

			safe = safe_sheet_name( name )
			if safe in wb.sheetnames:
				del wb[ safe ]
			ws = wb.create_sheet( safe )
			ws.append( self.HEADERS_ROW )
			for r_idx , row in enumerate( matched[ :1000 ] , start=2 ):
				for c_idx , val in enumerate( row , start=1 ):
					cell = ws.cell( row=r_idx , column=c_idx )
					if isinstance( val , utils.Link ):
						cell.value = val.text
						cell.hyperlink = val.url
						cell.style = "Hyperlink"
					else:
						cell.value = val
			print( f"'{name}': {len(matched)} hits" )

		wb.save( self.xlsx_path )
		print( f"appended {len(searches)} search sheets -> {self.xlsx_path}" )