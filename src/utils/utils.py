import re
import yaml
import json
import pickle
import base64
import unicodedata
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font
from rapidfuzz import fuzz , process
import importlib.util

from ..openalex import search_helpers

FUZZ_TITLE_THRESHOLD = 92

def write_yaml( file_path , python_object ):
	with open( file_path , "w" , encoding="utf-8" ) as f:
		yaml.dump( python_object , f )

def read_yaml( file_path ):
	with open( file_path , encoding="utf-8" ) as f:
		return yaml.safe_load( f )

def write_json( file_path , python_object ):
	with open( file_path , "w" , encoding="utf-8" ) as f:
		json.dump( python_object , f , ensure_ascii=False , indent=4 )

def read_json( file_path ):
	with open( file_path , encoding="utf-8" ) as f:
		return json.load( f )

def write_pickle( file_path , python_object ):
	with open( file_path , "wb" ) as f:
		pickle.dump( python_object , f )

def read_pickle( file_path ):
	with open( file_path , "rb" ) as f:
		return pickle.load( f )

# Crossref's recommended DOI pattern (the "good enough for 99%" one)
_DOI_RE = re.compile( r'10\.\d{4,9}/[-._;()/:a-z0-9]+' , re.IGNORECASE )
_DOI_TRAIL_RE = re.compile(
	r'/(figures?|tables?|abstract|full|epdf|pdf|fullpdf|references?|metrics|supplementary|suppl|media)'
	r'(/.*)?$' ,
	re.IGNORECASE
)
def normalize_doi( raw ):
	"""Extract a clean DOI from anywhere inside a string.
	Returns None if no DOI-like substring is found."""
	if not raw:
		return None
	v = str( raw ).strip()
	v = v.replace( "https://doi.org/" , "" ).replace( "http://doi.org/" , "" ).strip()
	m = _DOI_RE.search( v )
	if not m:
		return None
	doi = m.group( 0 )
	# strip trailing publisher path junk (/FIGURES/6, /full, etc.)
	doi = _DOI_TRAIL_RE.sub( '' , doi )
	# strip stray trailing punctuation the greedy class can grab
	doi = doi.rstrip( './;:)' )
	return doi

def normalize_title( s ):
	if not s:
		return ""
	# unicode normalization
	s = unicodedata.normalize( "NFKD" , s )
	# lowercase
	s = s.lower()
	# replace punctuation with space
	s = re.sub( r"[^a-z0-9]+" , " " , s )
	# collapse whitespace
	s = " ".join( s.split() )
	return s.strip()

def openalex_normalize_title( title ):
	if not title:
		return ""

	title = unicodedata.normalize( "NFKD" , str( title ) )
	title = title.replace( "|" , " " )
	title = title.replace( "“" , '"' ).replace( "”" , '"' )
	title = title.replace( "‘" , "'" ).replace( "’" , "'" )

	# Remove chars that OpenAlex search parser may treat as syntax/operators.
	title = re.sub( r"[|(){}\[\]^~?:\\/]" , " " , title )

	# Collapse whitespace.
	title = re.sub( r"\s+" , " " , title ).strip()

	return title

def base64_encode( message ):
	try:
		message_bytes = message.encode( 'utf-8' )
		base64_bytes = base64.b64encode( message_bytes )
		base64_message = base64_bytes.decode( 'utf-8' )
		return base64_message
	except Exception as e:
		print( e )
		return False

def base64_decode( base64_message ):
	try:
		base64_bytes = base64_message.encode( 'utf-8' )
		message_bytes = base64.b64decode(base64_bytes)
		message = message_bytes.decode( 'utf-8' )
		return message
	except Exception as e:
		print( e )
		return False

class Link:
	__slots__ = ( "text" , "url" )
	def __init__( self , text , url ):
		self.text , self.url = text , url
def write_xlsx( filepath , sheets ):
	wb = Workbook()
	wb.remove( wb.active )
	bold = Font( bold=True )
	for name , headers , rows in sheets:
		ws = wb.create_sheet( name )
		ws.append( list( headers ) )
		for c in ws[ 1 ]:
			c.font = bold
		for row in tqdm( list( rows ) , desc=f"Writing {name}" ):
			row = list( row )
			ws.append([ ( c.text if isinstance( c , Link ) else c ) for c in row ])
			for col_idx , c in enumerate( row , 1 ):
				if isinstance( c , Link ) and c.url:
					cell = ws.cell( row=ws.max_row , column=col_idx )
					cell.hyperlink = c.url
					cell.style = "Hyperlink"
	wb.save( filepath )

def title_of( d ):
	return d.get( "title" ) or d.get( "display_name" ) or ""
def openalex_to_xlsx_row( wid , meta , cite_count ):
	rd = meta.get( "doi" )
	clean_doi = normalize_doi( rd ) if rd else None
	proxy_url = f"https://doi-org.ezproxy.libraries.wright.edu/{clean_doi}" if clean_doi else None
	doi_url   = f"https://doi.org/{clean_doi}" if clean_doi else None
	proxy = Link( proxy_url , proxy_url ) if proxy_url else ""
	link  = Link( doi_url , doi_url ) if doi_url else ""
	best_oa = meta.get( "best_oa_location" ) or {}
	oa_block = meta.get( "open_access" ) or {}
	pdf_url = best_oa.get( "pdf_url" ) or oa_block.get( "oa_url" )
	pdf = Link( pdf_url , pdf_url ) if pdf_url else ""
	return [
		cite_count,
		title_of( meta ) or "(no metadata)" ,
		meta.get( "publication_year" ) ,
		proxy ,
		clean_doi ,
		link ,
		pdf ,
		meta.get( "cited_by_count" ) ,
		wid ,
	]

def build_library_dedup_index( snapshot ):
	"""Build a dedup index from a library snapshot ( normalized DOIs + titles ).
	Use with is_library_dup() to test whether an OpenAlex meta dict matches
	something already in the library."""
	dois = { normalize_doi( i[ "doi" ] ) for i in snapshot.values() if i.get( "doi" ) }
	dois.discard( None )
	titles_list = [
		normalize_title( i[ "title" ] )
		for i in snapshot.values() if i.get( "title" )
	]
	titles_list = [ t for t in titles_list if t ]
	titles_set = set( titles_list )
	return {
		"dois": dois ,
		"titles_set": titles_set ,
		"titles_list": titles_list ,
	}

def is_library_dup( meta , lib_index , fuzzy=True , threshold=FUZZ_TITLE_THRESHOLD ):
	"""True if meta matches the library by DOI ( exact ) or title.
	Title check: exact normalized first ; if fuzzy=True , also a
	token_set_ratio match at >= threshold. token_set_ratio is bag-of-words
	and false-positives on topical overlap , so pass fuzzy=False for
	high-precision callers ( e.g. the crawler , which feeds a manual review )."""
	rd = meta.get( "doi" )
	rd_norm = normalize_doi( rd ) if rd else None
	if rd_norm and rd_norm in lib_index[ "dois" ]:
		return True
	rt = meta.get( "title" ) or meta.get( "display_name" )
	rt_norm = normalize_title( rt ) if rt else None
	if not rt_norm:
		return False
	if rt_norm in lib_index[ "titles_set" ]:
		return True
	if not fuzzy:
		return False
	match = process.extractOne(
		rt_norm , lib_index[ "titles_list" ] ,
		scorer=fuzz.token_set_ratio ,
		score_cutoff=threshold ,
	)
	return match is not None

def load_searches( searches_dir , files=None ):
    """Load search predicates from .py files in searches_dir.
    If files is provided ( a list of names/paths ) , load only those instead of
    globbing the directory. Names without .py get the extension appended ;
    relative paths resolve against searches_dir ; absolute paths pass through."""
    from pathlib import Path
    if files:
        py_files = []
        for f in files:
            p = Path( f )
            if p.suffix != ".py":
                p = p.with_suffix( ".py" )
            if not p.is_absolute():
                p = searches_dir.joinpath( p )
            if not p.exists():
                raise FileNotFoundError( f"search file not found: {p}" )
            py_files.append( p )
    else:
        py_files = sorted( searches_dir.glob( "*.py" ) )
    all_searches = []
    for py_file in py_files:
        spec = importlib.util.spec_from_file_location( py_file.stem , py_file )
        module = importlib.util.module_from_spec( spec )
        for name in dir( search_helpers ):
            if not name.startswith( "__" ):
                setattr( module , name , getattr( search_helpers , name ) )
        spec.loader.exec_module( module )
        if hasattr( module , "search" ):
            all_searches.extend( module.search() )
    return all_searches