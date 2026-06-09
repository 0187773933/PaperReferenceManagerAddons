"""
prma search : run every predicate from args.searches / *.py against the
text we already have for each paper in db/papers/ and emit per-search
hits to
  args.output / searches / {search-slug}.md      ( one markdown per search )
  args.output / searches / searches.xlsx         ( one sheet  per search )

Same predicates the OpenAlex pipeline uses for the missing.xlsx search
sheets ( utils.load_searches ) -- this task just reuses them against the
text WE'VE produced for each library paper instead of against OpenAlex
abstracts.

Per-paper haystack ( lowercased before predicates see it ) :
  - paper title
  - paper[ 'raw_text' ]                                 ( pymupdf full text )
  - output/text/{doi}.txt                               ( ` prma text ` output )
  - output/methods/{doi}.txt                            ( ` prma methods ` output )
  - output/summaries/<section>/{doi}.md                 ( ` prma summarize ` LLM output )
Missing files are silently skipped -- whatever text you've generated so
far gets searched. Run more upstream tasks ( md , text , methods ,
summarize ) for a richer haystack.

Per-paper enrichment ( used for sorting + the md / xlsx columns ) :
  - paper[ 'pdf_path' ]                                 ( local PDF link )
  - output/cache/openalex/<base64(doi)>.json            ( publication_date , journal )

A predicate that throws is skipped for that paper ( logged once per
predicate ) so a buggy search doesn't take down a 2000-paper run.

Output files are rebuilt from scratch every invocation , so removing a
predicate from a search file makes its stale .md disappear next run.
"""

import re
from pathlib import Path
from collections import Counter

from tqdm import tqdm

from openpyxl import Workbook
from openpyxl.styles import Alignment , Font , PatternFill
from openpyxl.utils import get_column_letter

from ..db    import papers as papers_db
from ..utils import utils


# Same ezproxy host the rollup uses ; swap here if you fork for a
# different institution.
_PROXY_URL_TEMPLATE = "https://doi-org.ezproxy.libraries.wright.edu/{doi}"

# Header for both the markdown front-matter and xlsx column ordering.
_XLSX_COLUMNS = (
	( "DOI"        , 32 ) ,
	( "Title"      , 55 ) ,
	( "Published"  , 13 ) ,
	( "Journal"    , 30 ) ,
	( "Hashtags"   , 35 ) ,
	( "PDF"        , 40 ) ,
	( "Proxy"      , 50 ) ,
)

# openpyxl rejects ASCII control characters ; LLM / OCR output is
# usually clean , but belt + suspenders.
_ILLEGAL_XLSX_RE = re.compile( r"[\x00-\x08\x0B\x0C\x0E-\x1F]" )

_EXCEL_CELL_MAX = 32_000


# ---------------------------------------------------------------------------
# Helpers ( duplicated rather than imported from summarize_rollup so this
# task stays standalone -- they're 5-10 lines each )
# ---------------------------------------------------------------------------

def _scrub( text ):
	if not text:
		return text
	text = _ILLEGAL_XLSX_RE.sub( "" , text )
	if len( text ) > _EXCEL_CELL_MAX:
		text = text[ : _EXCEL_CELL_MAX ] + "\n[ ... truncated for Excel ... ]"
	return text


def _slug_for_filename( name ):
	"""'fMRI + inner speech' -> 'fMRI-inner-speech' . Keeps case for
	human readability ; collapses ALL whitespace + punctuation runs to a
	single '-' so different search names produce distinct files."""
	s = re.sub( r"[^A-Za-z0-9._-]+" , "-" , ( name or "" ).strip() )
	return s.strip( "-" ) or "search"


def _load_openalex_meta( args , doi ):
	"""Read output/cache/openalex/{base64( doi )}.json if present.
	Returns the parsed dict or {} on any miss."""
	if args is None or not doi:
		return {}
	b64 = utils.base64_encode( doi )
	if not b64:
		return {}
	fp = args.output.joinpath( "cache" , "openalex" , f"{b64}.json" )
	if not fp.exists():
		return {}
	try:
		return utils.read_json( fp ) or {}
	except Exception:
		return {}


def _openalex_journal_name( oa_meta ):
	if not oa_meta:
		return ""
	prim = ( oa_meta.get( "primary_location" ) or {} )
	src  = ( prim.get( "source" ) or {} )
	name = src.get( "display_name" )
	if name:
		return name
	hv = oa_meta.get( "host_venue" ) or {}
	return hv.get( "display_name" ) or ""


def _openalex_publication_date( oa_meta ):
	if not oa_meta:
		return ""
	pd = oa_meta.get( "publication_date" )
	if pd:
		return str( pd )
	py = oa_meta.get( "publication_year" )
	if py:
		return f"{py}-01-01"
	return ""


def _file_url( abs_path ):
	if not abs_path:
		return ""
	from urllib.parse import quote
	return "file://" + quote( str( abs_path ) , safe="/:" )


# ---------------------------------------------------------------------------
# Haystack assembly + hashtag harvest
# ---------------------------------------------------------------------------

# Match the per-paper summary md's "**Hashtags:** ..." line ( same shape
# as src/tasks/summarize._render_md emits ).
_HASHTAGS_LINE_RE = re.compile(
	r"^\*\*Hashtags:\*\*\s+(.+?)\s*$" , re.MULTILINE ,
)


def _summary_md_paths( args , doi ):
	"""Yield Path objects for every existing per-section LLM summary
	md file for this paper."""
	prefix = utils.doi_to_filename( doi )
	if not prefix:
		return
	root = args.output.joinpath( "summaries" )
	if not root.exists():
		return
	for section_dir in sorted( root.iterdir() ):
		if not section_dir.is_dir():
			continue
		f = section_dir.joinpath( f"{prefix}.md" )
		if f.exists():
			yield f


def _per_section_text_paths( args , doi ):
	"""Yield Path objects for ` prma text ` , ` prma methods ` , and any
	other per-paper-text-per-section .txt files we might add later."""
	prefix = utils.doi_to_filename( doi )
	if not prefix:
		return
	for subdir in ( "text" , "methods" ):
		f = args.output.joinpath( subdir , f"{prefix}.txt" )
		if f.exists():
			yield f


def _build_haystack_and_hashtags( args , doi , paper ):
	"""Stitch together every text source we have for this paper and
	return ( haystack_lower , hashtags_str ). The haystack is lowercased
	for predicate consumption ; hashtags is a space-joined union of
	every '**Hashtags:**' line found across the per-section summary
	mds."""
	parts = []
	hashtags = []

	title = ( paper.get( "title" ) or "" ).strip()
	if title:
		parts.append( title )

	raw = paper.get( "raw_text" ) or ""
	if raw:
		parts.append( raw )

	for f in _summary_md_paths( args , doi ):
		try:
			text = f.read_text( encoding="utf-8" )
		except Exception:
			continue
		parts.append( text )
		m = _HASHTAGS_LINE_RE.search( text )
		if m:
			for tok in m.group( 1 ).split():
				if tok.startswith( "#" ) and len( tok ) > 1:
					hashtags.append( tok )

	for f in _per_section_text_paths( args , doi ):
		try:
			parts.append( f.read_text( encoding="utf-8" ) )
		except Exception:
			continue

	haystack = "\n\n".join( parts ).lower()
	# Dedupe hashtags while preserving first-seen order.
	seen = set()
	dedup_tags = []
	for t in hashtags:
		if t not in seen:
			seen.add( t )
			dedup_tags.append( t )
	return haystack , " ".join( dedup_tags )


# ---------------------------------------------------------------------------
# Per-search markdown emitter
# ---------------------------------------------------------------------------

def _render_search_md( matches , search_name , out_path , total_n ):
	"""Emit one markdown list of papers that matched a single search.
	`matches` is a list of row-dicts already enriched with openalex /
	pdf_path / hashtags. Sorts by publication_date desc , missing dates
	to the end. Empty match list -> delete any stale .md and return 0."""
	if not matches:
		if out_path.exists():
			try: out_path.unlink()
			except Exception: pass
		return 0

	def _key( m ):
		return _openalex_publication_date( m.get( "openalex" ) or {} ) or "0000-00-00"
	matches = sorted( matches , key=_key , reverse=True )

	lines = []
	lines.append( f"# {search_name} — local library search" )
	lines.append( "" )
	lines.append(
		f"_{ len( matches ) } of { total_n } papers matched. Sorted by "
		f"publication date ( most recent first ). Generated by `prma search`._"
	)
	lines.append( "" )
	lines.append( "---" )
	lines.append( "" )

	for m in matches:
		title    = ( m.get( "title" ) or "(untitled)" ).strip()
		doi      = m.get( "doi" )    or ""
		pdf_path = m.get( "pdf_path" )
		oa       = m.get( "openalex" ) or {}
		pub_date = _openalex_publication_date( oa ) or "(unknown)"
		journal  = _openalex_journal_name    ( oa ) or "(unknown)"
		hashtags = m.get( "hashtags" ) or ""
		summary_links = m.get( "summary_links" ) or []

		lines.append( f"## {title}" )
		lines.append( "" )
		if doi:
			lines.append( f"- **DOI:** [{doi}](https://doi.org/{doi})" )
			proxy = _PROXY_URL_TEMPLATE.format( doi=doi )
			lines.append( f"- **Proxy:** [{proxy}]({proxy})" )
		else:
			lines.append( "- **DOI:** _(none)_" )
		if pdf_path:
			lines.append(
				f"- **PDF:** [{Path( pdf_path ).name}]({_file_url( pdf_path )})"
			)
		else:
			lines.append( "- **PDF:** _(no local pdf)_" )
		lines.append( f"- **Published:** {pub_date}" )
		lines.append( f"- **Journal:** {journal}" )
		if hashtags:
			lines.append( f"- **Hashtags:** {hashtags}" )
		if summary_links:
			lines.append( "- **Section summaries:**" )
			for section , path in summary_links:
				lines.append( f"  - [{section}]({_file_url( path )})" )
		lines.append( "" )
		lines.append( "---" )
		lines.append( "" )

	out_path.parent.mkdir( parents=True , exist_ok=True )
	out_path.write_text( "\n".join( lines ) , encoding="utf-8" )
	return len( matches )


# ---------------------------------------------------------------------------
# Combined xlsx ( one sheet per search )
# ---------------------------------------------------------------------------

def _safe_sheet_name( name ):
	"""Excel hard-caps sheet names at 31 chars and bans :/\\?*[] ."""
	safe = name
	for ch in ":/\\?*[]":
		safe = safe.replace( ch , " " )
	safe = re.sub( r"\s+" , " " , safe ).strip()
	return safe[ :31 ] or "search"


def _write_workbook( out_path , results ):
	"""One sheet per search. `results` is a dict mapping search_name ->
	list of match dicts ( already enriched ). Same column order across
	every sheet ( see _XLSX_COLUMNS )."""
	wb = Workbook()
	# Drop the implicit default sheet ; we'll create our own.
	wb.remove( wb.active )

	header_font = Font( bold=True , color="FFFFFF" )
	header_fill = PatternFill( "solid" , fgColor="046A38" )
	header_align = Alignment( horizontal="center" , vertical="center" , wrap_text=True )
	wrap_top    = Alignment( wrap_text=True , vertical="top" )
	link_top    = Alignment( wrap_text=False , vertical="top" )

	used_sheet_names = set()
	for search_name , matches in results.items():
		# Skip empty searches -- no need for blank sheets.
		if not matches:
			continue

		# Sort by publication_date desc.
		def _key( m ):
			return _openalex_publication_date( m.get( "openalex" ) or {} ) or "0000-00-00"
		matches = sorted( matches , key=_key , reverse=True )

		base = _safe_sheet_name( search_name )
		name , n = base , 2
		while name in used_sheet_names:
			suffix = f" ({n})"
			name = base[ : 31 - len( suffix ) ] + suffix
			n += 1
		used_sheet_names.add( name )

		ws = wb.create_sheet( name )
		ws.append( [ h for h , _ in _XLSX_COLUMNS ] )
		for cell in ws[ 1 ]:
			cell.font      = header_font
			cell.fill      = header_fill
			cell.alignment = header_align
		for i , ( _ , width ) in enumerate( _XLSX_COLUMNS , 1 ):
			ws.column_dimensions[ get_column_letter( i ) ].width = width

		for m in matches:
			doi      = m.get( "doi" ) or ""
			pdf_path = m.get( "pdf_path" )
			oa       = m.get( "openalex" ) or {}
			ws.append( [
				_scrub( doi ) ,
				_scrub( m.get( "title" ) or "" ) ,
				_openalex_publication_date( oa ) or "" ,
				_scrub( _openalex_journal_name    ( oa ) ) ,
				_scrub( m.get( "hashtags" ) or "" ) ,
				_scrub( Path( pdf_path ).name if pdf_path else "" ) ,
				_PROXY_URL_TEMPLATE.format( doi=doi ) if doi else "" ,
			] )
			r = ws.max_row

			# DOI hyperlink.
			c_doi = ws.cell( row=r , column=1 )
			if doi:
				c_doi.hyperlink = f"https://doi.org/{doi}"
				c_doi.style     = "Hyperlink"
			c_doi.alignment = link_top

			# Title wraps.
			ws.cell( row=r , column=2 ).alignment = wrap_top

			# PDF column : hyperlink to the local file:// URL when we have one.
			c_pdf = ws.cell( row=r , column=6 )
			if pdf_path:
				c_pdf.hyperlink = _file_url( pdf_path )
				c_pdf.style     = "Hyperlink"
			c_pdf.alignment = link_top

			# Proxy column : straight hyperlink.
			c_proxy = ws.cell( row=r , column=7 )
			if doi:
				c_proxy.hyperlink = _PROXY_URL_TEMPLATE.format( doi=doi )
				c_proxy.style     = "Hyperlink"
			c_proxy.alignment = link_top

			# Hashtags wraps so a wide tag list reads well.
			ws.cell( row=r , column=5 ).alignment = wrap_top

		ws.freeze_panes = "B2"
		last_col = get_column_letter( len( _XLSX_COLUMNS ) )
		ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

	# If every search came up empty , wb has no sheets and save() blows
	# up. Add a placeholder.
	if not wb.sheetnames:
		ws = wb.create_sheet( "no matches" )
		ws[ "A1" ] = "No papers matched any search predicate."
	wb.save( out_path )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def _resolve_managers( args ):
	name = ( getattr( args , "manager" , None ) or "zotero" ).lower()
	if getattr( args , "mendeley" , False ):
		name = "mendeley"
	elif getattr( args , "zotero" , False ):
		name = "zotero"
	if name == "all":
		return ( papers_db.SOURCE_ZOTERO , papers_db.SOURCE_MENDELEY )
	return ( name , )


def _paper_matches_managers( paper , managers ):
	if not managers:
		return True
	srcs = paper.get( "sources" ) or {}
	return any( m in srcs for m in managers )


def run( args ):
	# Load search predicates ( same loader the OpenAlex pipeline uses ).
	searches = utils.load_searches(
		args.searches ,
		files=getattr( args , "search_files" , None ) ,
	)
	if not searches:
		print(
			f"SEARCH    :: no predicates found under {args.searches} "
			f"( drop .py files there ; see existing examples )"
		)
		return

	managers = _resolve_managers( args )
	manager_label = " + ".join( managers ) if managers else "all"

	out_dir = args.output.joinpath( "searches" )
	out_dir.mkdir( parents=True , exist_ok=True )

	print(
		f"SEARCH    :: ({manager_label}) running {len( searches )} predicate(s) "
		f"across the unified library -> {out_dir}"
	)

	# Collect matches per search.
	results = { name: [] for name , _ in searches }
	broken_predicates = set()
	n_total = 0

	for doi , paper in tqdm(
		list( papers_db.iter_all( args ) ) ,
		desc="papers" , unit="paper" ,
	):
		if not _paper_matches_managers( paper , managers ):
			continue
		n_total += 1
		haystack , hashtags = _build_haystack_and_hashtags( args , doi , paper )
		if not haystack:
			continue
		# Pre-build the match dict ; reused if multiple searches hit.
		match_proto = None
		for search_name , predicate in searches:
			if search_name in broken_predicates:
				continue
			try:
				hit = predicate( haystack )
			except Exception as e:
				broken_predicates.add( search_name )
				print(
					f"SEARCH    :: predicate {search_name!r} threw "
					f"( {e} ) ; skipping for the rest of the run"
				)
				continue
			if not hit:
				continue
			if match_proto is None:
				# Hydrate openalex meta + summary-link list only once per
				# matching paper ; cheap if there are many predicates.
				summary_links = []
				prefix = utils.doi_to_filename( doi )
				if prefix:
					summaries_root = args.output.joinpath( "summaries" )
					if summaries_root.exists():
						for section_dir in sorted( summaries_root.iterdir() ):
							if not section_dir.is_dir():
								continue
							f = section_dir.joinpath( f"{prefix}.md" )
							if f.exists():
								summary_links.append( ( section_dir.name , f ) )
				match_proto = {
					"doi"           : doi ,
					"title"         : paper.get( "title" ) ,
					"pdf_path"      : paper.get( "pdf_path" ) ,
					"hashtags"      : hashtags ,
					"openalex"      : _load_openalex_meta( args , doi ) ,
					"summary_links" : summary_links ,
				}
			results[ search_name ].append( match_proto )

	# Emit per-search markdown.
	for search_name , matches in results.items():
		slug    = _slug_for_filename( search_name )
		md_path = out_dir.joinpath( f"{slug}.md" )
		n = _render_search_md( matches , search_name , md_path , n_total )
		if n:
			print( f"SEARCH    :: {search_name!r}: {n} matches -> {md_path}" )
		else:
			print( f"SEARCH    :: {search_name!r}: 0 matches" )

	# Emit combined workbook.
	xlsx_path = out_dir.joinpath( "searches.xlsx" )
	_write_workbook( xlsx_path , results )
	print( f"SEARCH    :: combined workbook -> {xlsx_path}" )
