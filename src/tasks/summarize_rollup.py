"""
Per-section rollup ( ` prma rollup [section] ` ) : scan every summary .md
the LLM has written under
  args.output / summaries / {section} /
and aggregate them into a single Excel workbook at
  args.output / summaries / {section}.xlsx

Run AFTER ` prma summarize ` has produced the per-paper markdown files.
Decoupled into its own command ( instead of an auto-step at the end of
summarize ) so the rollup is cheap to re-run / debug : tweaks to the
xlsx layout or to the subsection parser can be exercised without paying
for another round of LLM calls.

The 'creative heuristic' is to mine the LLM's OWN structured output -- our
prompts ( config/llm-prompts/*.txt ) tell the model to emit a fixed set
of "**Subsection.**" bold headers ( e.g. for methods :
  **Study design.** , **Participants.** , **Materials & equipment.** ,
  ... ) -- and turn each of those headers into its own COLUMN. Different
papers can have different subsets of subsections present ; we keep the
union of headers seen across all papers , in first-appearance order
( alphabetical-by-DOI ) , so the column layout is deterministic and
covers everything the LLM emitted.

Two sheets :

  "Summaries"      -- one row per paper :
      DOI ( hyperlink ) , Title , Hashtags ,
      < one column per discovered subsection > , Full Summary

      Wrap-text + sensible widths + frozen header + autofilter. Reading
      across one row is the per-paper executive summary ; reading down
      a column is "every paper's Methods.Statistics" or "every paper's
      Results.Primary outcome" at a glance.

  "Hashtag counts" -- the rolled-up frequency table of every hashtag the
      LLM emitted across the cohort , sorted descending. Useful as a
      seed for clustering or for spotting outlier tags worth pinning.

Plus , for every extra column declared in config/rollup-extra-columns.yaml
( e.g. "fMRI task" ) , a sibling markdown list at
  args.output / summaries / {section}-{column-slug}.md
listing every paper whose snippets matched , sorted by OpenAlex
publication_date ( most recent first ). Each entry carries title , DOI ,
ezproxy URL , local PDF link , publication date , journal , and the
snippet text from both the LLM summary and the raw section text.

Idempotent : every run rebuilds the xlsx ( and every extra-column .md )
from the current set of .md files under the section folder , so adding /
re-running individual papers and then a rollup pass produces an up-to-
date workbook + lists.
"""

import re
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment , Font , PatternFill
from openpyxl.utils import get_column_letter

from ..utils import utils


# ---------------------------------------------------------------------------
# Per-md parser
# ---------------------------------------------------------------------------

# Matches the H1 we emit : "# {SectionDisplay} — {Title}" . The em-dash
# is U+2014 ; we also accept a plain " - " for tolerance in case
# someone hand-edits a file.
_H1_RE        = re.compile( r"^#\s+(.+?)(?:\s+[—\-]\s+(.+))?\s*$" , re.MULTILINE )
_DOI_RE       = re.compile( r"^\*\*DOI:\*\*\s+\[([^\]]+)\]\([^)]+\)"   , re.MULTILINE )
_HASHTAGS_RE  = re.compile( r"^\*\*Hashtags:\*\*\s+(.+?)\s*$"          , re.MULTILINE )
_SUMMARY_H2_RE = re.compile( r"^##\s+Summary\s*$"                       , re.MULTILINE )

# A subsection header is a line that's ENTIRELY a single bolded label :
#   "**Study design.**" , "**Mechanism / interpretation.**" , "**Primary outcome.**"
# Length-capped so inline bold like "**important**" anywhere in body
# prose doesn't get mistaken for a header. The trailing period is
# optional ( the LLM occasionally drops it ) , and we strip any
# trailing colon ( the prompt formats lean toward "Header." but
# "Header:" creeps in ).
_SUBSECTION_HEADER_RE = re.compile(
	r"^\*\*([^*\n]{1,80}?)\*\*\s*$" ,
	re.MULTILINE ,
)


def _split_subsections( body ):
	"""Return [ ( header , content ) ... ] split on the bold one-line
	subsection headers. Headers are normalized ( trailing period / colon
	dropped ; whitespace collapsed ). If no headers are found , the
	whole body is returned under the synthetic key '(body)'."""
	matches = list( _SUBSECTION_HEADER_RE.finditer( body ) )
	if not matches:
		stripped = body.strip()
		return [ ( "(body)" , stripped ) ] if stripped else []
	out = []
	for i , m in enumerate( matches ):
		header = m.group( 1 ).strip().rstrip( ".:" ).strip()
		header = re.sub( r"\s+" , " " , header )
		start  = m.end()
		end    = matches[ i + 1 ].start() if i + 1 < len( matches ) else len( body )
		content = body[ start : end ].strip()
		if header:
			out.append( ( header , content ) )
	return out


def _parse_summary_md( text , fallback_doi="" ):
	"""Pull doi , title , hashtags , and an ordered list of
	( subsection_header , content ) pairs out of one summary .md file.
	Missing fields come back as empty strings ; the caller can decide
	whether to skip the row."""
	doi      = ""
	title    = ""
	hashtags = ""
	subsections = []
	full_summary = ""

	m_doi = _DOI_RE.search( text )
	if m_doi:
		doi = m_doi.group( 1 ).strip()
	if not doi:
		doi = fallback_doi

	m_h1 = _H1_RE.search( text )
	if m_h1:
		title = ( m_h1.group( 2 ) or m_h1.group( 1 ) or "" ).strip()

	m_tags = _HASHTAGS_RE.search( text )
	if m_tags:
		hashtags = m_tags.group( 1 ).strip()

	m_sum = _SUMMARY_H2_RE.search( text )
	if m_sum:
		full_summary = text[ m_sum.end() : ].strip()
		subsections  = _split_subsections( full_summary )

	return {
		"doi"          : doi          ,
		"title"        : title        ,
		"hashtags"     : hashtags     ,
		"subsections"  : subsections  ,
		"full_summary" : full_summary ,
	}


# ---------------------------------------------------------------------------
# Extra columns ( config-driven keyword / regex extraction )
# ---------------------------------------------------------------------------

EXTRA_COLUMNS_FILENAME = "rollup-extra-columns.yaml"


def _load_extra_columns_config( config_dir ):
	"""Read config/rollup-extra-columns.yaml. Returns the list of column
	specs , or [] if the file is missing or empty. A malformed file logs
	once and returns []."""
	if config_dir is None:
		return []
	path = config_dir.joinpath( EXTRA_COLUMNS_FILENAME )
	if not path.exists():
		return []
	try:
		data = utils.read_yaml( path ) or {}
	except Exception as e:
		print( f"ROLLUP    :: could not read {path} ( {e} ) ; "
		       f"skipping extra columns" )
		return []
	cols = data.get( "extra_columns" ) or []
	cleaned = []
	for raw in cols:
		if not isinstance( raw , dict ):
			continue
		name = ( raw.get( "name" ) or "" ).strip()
		if not name:
			continue
		spec = {
			"name"         : name ,
			"keywords"     : [ str( k ) for k in ( raw.get( "keywords" ) or [] ) ] ,
			"regex"        : [ str( r ) for r in ( raw.get( "regex"    ) or [] ) ] ,
			"sources"      : list( raw.get( "sources" ) or ( "summary" , "section_text" ) ) ,
			"context_chars": int( raw.get( "context_chars" , 300 ) ) ,
			"max_snippets" : int( raw.get( "max_snippets"  , 4   ) ) ,
		}
		# Pre-compile patterns once at load time.
		spec[ "keyword_re" ] = _build_keyword_regex( spec[ "keywords" ] )
		spec[ "regex_re" ]   = _build_regex_list  ( spec[ "regex"   ] , spec[ "name" ] )
		cleaned.append( spec )
	return cleaned


def _build_keyword_regex( keywords ):
	"""OR-combine the keywords into one case-insensitive regex with
	non-word-character lookaround boundaries -- so multi-word phrases
	like "fMRI task" match , and single-word keywords like "scanner"
	don't fire mid-word ( "subscanner" doesn't count )."""
	parts = []
	for k in keywords:
		k = ( k or "" ).strip()
		if not k:
			continue
		prefix = r"(?<!\w)" if k[ 0  ].isalnum() else ""
		suffix = r"(?!\w)"  if k[ -1 ].isalnum() else ""
		parts.append( prefix + re.escape( k ) + suffix )
	if not parts:
		return None
	return re.compile( "|".join( parts ) , re.IGNORECASE )


def _build_regex_list( patterns , col_name ):
	"""Compile each user-supplied regex. A bad pattern logs once and is
	dropped ; the rest still work."""
	out = []
	for p in patterns:
		p = ( p or "" ).strip()
		if not p:
			continue
		try:
			out.append( re.compile( p , re.IGNORECASE ) )
		except re.error as e:
			print(
				f"ROLLUP    :: extra-column {col_name!r} regex {p!r} "
				f"failed to compile ( {e} ) ; dropped"
			)
	return out


def _snap_to_sentence( text , start , end ):
	"""Widen [ start , end ) outward to the nearest sentence boundary on
	each side ( previous '.' / '!' / '?' / '\\n' going backward , next
	'.' / '!' / '?' / '\\n' going forward ). Keeps snippets readable in
	the xlsx cell."""
	# Backward : prefer paragraph break , else sentence end.
	back_para = text.rfind( "\n\n" , 0 , start )
	back_sent = max(
		text.rfind( ". "  , 0 , start ) ,
		text.rfind( "! "  , 0 , start ) ,
		text.rfind( "? "  , 0 , start ) ,
		text.rfind( "\n"  , 0 , start ) ,
	)
	new_start = max( back_para + 2 if back_para >= 0 else 0 ,
	                 back_sent + 2 if back_sent >= 0 else 0 )
	if new_start < 0:
		new_start = 0

	# Forward : end at the next sentence terminator inside the window
	# or the next paragraph break.
	candidates = []
	for sep in ( ". " , "! " , "? " , "\n\n" , "\n" ):
		idx = text.find( sep , end )
		if idx >= 0:
			candidates.append( idx + len( sep ) )
	new_end = min( candidates ) if candidates else len( text )
	new_end = min( new_end , len( text ) )

	return new_start , new_end


def _extract_snippets( text , spec ):
	"""Return up to spec[ 'max_snippets' ] de-duplicated snippets where
	any spec keyword OR regex matches in `text` , each snapped to
	sentence boundaries and roughly spec[ 'context_chars' ] long."""
	if not text:
		return []
	hits = []
	rx_kw = spec.get( "keyword_re" )
	if rx_kw is not None:
		hits.extend( ( m.start() , m.end() ) for m in rx_kw.finditer( text ) )
	for rx in spec.get( "regex_re" , [] ):
		hits.extend( ( m.start() , m.end() ) for m in rx.finditer( text ) )
	if not hits:
		return []
	hits.sort()

	max_snips = max( 1 , spec.get( "max_snippets" , 4 ) )
	ctx       = max( 80 , spec.get( "context_chars" , 300 ) )
	half      = ctx // 2

	snippets  = []
	covered   = []        # list of ( start , end ) ranges already emitted
	for h_start , h_end in hits:
		# Initial window centered on the match.
		win_start = max( 0 , h_start - half )
		win_end   = min( len( text ) , h_end + half )
		# Snap to sentence boundaries for readability.
		s , e = _snap_to_sentence( text , win_start , win_end )
		# Skip if the new range overlaps a previously-emitted one ; the
		# hits are already sorted , so we only need to check the last.
		if covered and s < covered[ -1 ][ 1 ]:
			continue
		snippet = text[ s : e ].strip()
		if not snippet:
			continue
		snippets.append( snippet )
		covered.append( ( s , e ) )
		if len( snippets ) >= max_snips:
			break
	return snippets


def _gather_extras_for_row( row , section_key , args , extra_columns ):
	"""For one paper , run every extra-column spec across its sources
	and stuff the joined snippet into row[ 'extras' ][ col_name ]. The
	per-source snippets are kept separately under row[ 'extras_by_src' ]
	[ col_name ] = { 'summary': [ str , ... ] , 'raw': [ str , ... ] }
	so the per-column markdown emitter ( _render_extra_column_md ) can
	cite source provenance cleanly. The xlsx writer keeps using the
	flat row[ 'extras' ] string."""
	if not extra_columns:
		return
	# Pre-resolve the two text sources once per paper.
	summary_text = row.get( "full_summary" ) or ""
	section_text = ""
	doi = row.get( "doi" ) or ""
	if doi and args is not None:
		try:
			from ..pdf import section_text as ST
			section_text = ST.resolve_section_text( args , doi , section_key )
		except Exception:
			section_text = ""
	sources_lookup = {
		"summary"      : summary_text ,
		"section_text" : section_text ,
	}

	extras        = {}
	extras_by_src = {}
	for spec in extra_columns:
		all_snips   = []
		per_src     = { "summary": [] , "raw": [] }
		for src_name in spec[ "sources" ]:
			src_text = sources_lookup.get( src_name )
			if not src_text:
				continue
			tag = "summary" if src_name == "summary" else "raw"
			for snip in _extract_snippets( src_text , spec ):
				per_src[ tag ].append( snip )
				all_snips.append( f"[ {tag} ] {snip}" )
				if len( all_snips ) >= spec[ "max_snippets" ]:
					break
			if len( all_snips ) >= spec[ "max_snippets" ]:
				break
		extras       [ spec[ "name" ] ] = "\n\n---\n\n".join( all_snips )
		extras_by_src[ spec[ "name" ] ] = per_src
	row[ "extras" ]        = extras
	row[ "extras_by_src" ] = extras_by_src


# ---------------------------------------------------------------------------
# Per-extra-column Markdown emitter
# ---------------------------------------------------------------------------

# Proxy template for the Wright State ezproxy ; if you fork this for a
# different institution , swap the host here.
_PROXY_URL_TEMPLATE = "https://doi-org.ezproxy.libraries.wright.edu/{doi}"


def _load_openalex_meta( args , doi ):
	"""Read output/cache/openalex/{base64( doi )}.json if present.
	Returns the parsed dict or {} on any miss. Quiet : missing files
	are the normal case for papers whose openalex cache hasn't been
	fetched yet ( i.e. ` prma ` hasn't run since they were added )."""
	if args is None or not doi:
		return {}
	b64 = utils.base64_encode( doi )
	if not b64:
		return {}
	fp = args.output.joinpath( "cache" , "openalex" , f"{b64}.json" )
	if not fp.exists():
		return {}
	try:
		return utils.read_json( fp ) or {}
	except Exception:
		return {}


def _openalex_journal_name( oa_meta ):
	"""Pull a human-readable journal / venue name out of an OpenAlex
	Works record. Tries the current API field first
	( primary_location.source.display_name ) , then the older field
	( host_venue.display_name ) , then '' ."""
	if not oa_meta:
		return ""
	prim = ( oa_meta.get( "primary_location" ) or {} )
	src  = ( prim.get( "source" ) or {} )
	name = src.get( "display_name" )
	if name:
		return name
	hv = oa_meta.get( "host_venue" ) or {}
	return hv.get( "display_name" ) or ""


def _openalex_publication_date( oa_meta ):
	"""'YYYY-MM-DD' if OpenAlex has it , else just the year , else '' .
	Returns a string so it sorts correctly by lex order downstream."""
	if not oa_meta:
		return ""
	pd = oa_meta.get( "publication_date" )
	if pd:
		return str( pd )
	py = oa_meta.get( "publication_year" )
	if py:
		return f"{py}-01-01"
	return ""


def _file_url( abs_path ):
	"""Build a file:// URL from an absolute path , URL-escaping spaces
	and other meta characters so markdown viewers ( and shell pastes )
	don't choke."""
	if not abs_path:
		return ""
	from urllib.parse import quote
	# quote with no safe chars eats slashes too -- keep them so the URL
	# is recognizably a path.
	return "file://" + quote( str( abs_path ) , safe="/:" )


def _slug_for_filename( name ):
	"""Turn 'fMRI task' into 'fMRI-task' so we can use the column name
	as a filename component. Preserves letter case ( so the slug echoes
	the user's column name ) , just collapses whitespace / punctuation
	to '-' and trims leading / trailing dashes."""
	s = re.sub( r"[^A-Za-z0-9._-]+" , "-" , ( name or "" ).strip() )
	return s.strip( "-" ) or "extra"


def _md_escape_pipe( s ):
	"""Defensive : strip control chars and replace pipes in single-line
	cells. Used for title fields that go into bullet points."""
	if not s:
		return ""
	s = _ILLEGAL_XLSX_RE.sub( "" , s )
	return s.strip()


def _render_extra_column_md(
	rows , col_spec , section_key , section_display , out_path , total_n ,
):
	"""Emit a single markdown list of every paper whose extras[ col_name ]
	matched , sorted by OpenAlex publication_date ( most recent first ).
	Each entry gets : title , doi , proxy url , local pdf link , pub
	date , journal , and the per-source snippet text. Papers with no
	openalex meta still appear ( published / journal fields show as
	'(unknown)' ) and sort to the end of the list."""
	col_name = col_spec[ "name" ]

	# Filter to matching papers.
	matched = [ r for r in rows if ( r.get( "extras" ) or {} ).get( col_name ) ]
	if not matched:
		# Don't write an empty file -- delete a stale one if present , so
		# users don't get fooled by leftover content from a prior run.
		if out_path.exists():
			try:
				out_path.unlink()
			except Exception:
				pass
		return 0

	# Sort by publication_date desc ; missing dates sort to the end.
	def _key( r ):
		oa = r.get( "openalex" ) or {}
		pd = _openalex_publication_date( oa )
		return pd or "0000-00-00"
	matched.sort( key=_key , reverse=True )

	# Build the markdown.
	lines = []
	lines.append( f"# {col_name} — {section_display}" )
	lines.append( "" )
	lines.append(
		f"_{ len( matched ) } of { total_n } papers matched. Sorted by "
		f"publication date ( most recent first ). Generated by `prma rollup`._"
	)
	lines.append( "" )
	lines.append( "---" )
	lines.append( "" )

	for r in matched:
		title    = _md_escape_pipe( r.get( "title" ) or "(untitled)" )
		doi      = r.get( "doi" )    or ""
		pdf_path = r.get( "pdf_path" )
		oa_meta  = r.get( "openalex" ) or {}
		pub_date = _openalex_publication_date( oa_meta ) or "(unknown)"
		journal  = _openalex_journal_name    ( oa_meta ) or "(unknown)"

		lines.append( f"## {title}" )
		lines.append( "" )
		if doi:
			lines.append( f"- **DOI:** [{doi}](https://doi.org/{doi})" )
			proxy = _PROXY_URL_TEMPLATE.format( doi=doi )
			lines.append( f"- **Proxy:** [{proxy}]({proxy})" )
		else:
			lines.append( "- **DOI:** _(none)_" )
		if pdf_path:
			from pathlib import Path
			pdf_name = Path( pdf_path ).name
			lines.append( f"- **PDF:** [{pdf_name}]({_file_url( pdf_path )})" )
		else:
			lines.append( "- **PDF:** _(no local pdf)_" )
		lines.append( f"- **Published:** {pub_date}" )
		lines.append( f"- **Journal:** {journal}" )
		lines.append( "" )

		# Per-source snippets. Prefer the structured form built by
		# _gather_extras_for_row ; fall back to the flat string when
		# we're rendering an older row dict.
		by_src   = ( r.get( "extras_by_src" ) or {} ).get( col_name ) or {}
		sum_hits = by_src.get( "summary" ) or []
		raw_hits = by_src.get( "raw"     ) or []
		if sum_hits or raw_hits:
			lines.append( f"**{col_name} hits:**" )
			lines.append( "" )
			for s in sum_hits:
				lines.append( f"_[ from LLM summary ]_" )
				lines.append( "" )
				lines.append( s.strip() )
				lines.append( "" )
			for s in raw_hits:
				lines.append( f"_[ from raw {section_key} text ]_" )
				lines.append( "" )
				lines.append( s.strip() )
				lines.append( "" )
		else:
			# Fall back to the flat blob.
			lines.append( f"**{col_name} hits:**" )
			lines.append( "" )
			lines.append( r[ "extras" ][ col_name ].strip() )
			lines.append( "" )

		lines.append( "---" )
		lines.append( "" )

	out_path.parent.mkdir( parents=True , exist_ok=True )
	out_path.write_text( "\n".join( lines ) , encoding="utf-8" )
	return len( matched )


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

# openpyxl rejects ASCII control characters ( other than \t \n \r ) in
# cell values ; LLM output is generally clean but we belt-and-suspender.
_ILLEGAL_XLSX_RE = re.compile( r"[\x00-\x08\x0B\x0C\x0E-\x1F]" )

# Per-cell character cap : Excel hard-limits to 32767 chars.
_EXCEL_CELL_MAX = 32_000


def _scrub( text ):
	if not text:
		return text
	text = _ILLEGAL_XLSX_RE.sub( "" , text )
	if len( text ) > _EXCEL_CELL_MAX:
		text = text[ : _EXCEL_CELL_MAX ] + "\n[ ... truncated for Excel cell limit ... ]"
	return text


def _collect_rows( section_dir , args=None , section_key=None , extra_columns=None ):
	"""Walk every .md in section_dir ( alphabetical by filename , which
	is the doi_to_filename-encoded DOI ) and parse each. Returns
	  ( rows , ordered_headers , tag_counter )
	where ordered_headers is the union of subsection headers seen across
	all papers in first-appearance order. When extra_columns is non-empty
	each row also gets row[ 'extras' ] = { col_name -> snippet_text }."""
	md_files = sorted( section_dir.glob( "*.md" ) )
	rows             = []
	ordered_headers  = []
	tag_counter      = Counter()
	seen_headers     = set()
	for path in md_files:
		try:
			raw = path.read_text( encoding="utf-8" )
		except Exception as e:
			print( f"ROLLUP :: could not read {path} ( {e} )" )
			continue
		# Reverse the filename->DOI encoding for the fallback DOI.
		fallback_doi = path.stem.replace( "_" , "/" )
		parsed = _parse_summary_md( raw , fallback_doi=fallback_doi )
		# Skip entries that have neither a doi nor a body -- almost
		# certainly a corrupt / empty file.
		if not parsed[ "doi" ] and not parsed[ "full_summary" ]:
			continue
		# Index subsections by normalized header so wide-row lookup is
		# straightforward downstream.
		sub_map = {}
		for header , content in parsed[ "subsections" ]:
			if header not in seen_headers:
				ordered_headers.append( header )
				seen_headers.add( header )
			# If the same header appears twice in one paper ( rare ; LLM
			# echo ) , concatenate rather than drop the second hit.
			if header in sub_map:
				sub_map[ header ] = sub_map[ header ] + "\n\n" + content
			else:
				sub_map[ header ] = content
		parsed[ "sub_map" ] = sub_map
		# Config-driven keyword grep across summary + raw section text.
		# Done here ( inside the file walk ) so we keep the path open for
		# possible future per-file caching ; for now it just calls
		# resolve_section_text per paper.
		if extra_columns:
			_gather_extras_for_row( parsed , section_key , args , extra_columns )
			# Only papers that matched at least one extra column need
			# the per-paper enrichment ( pdf path + openalex meta ) for
			# the markdown emitter -- everyone else stays cheap.
			if any( ( parsed.get( "extras" ) or {} ).values() ) and parsed[ "doi" ] and args is not None:
				try:
					from ..db import papers as papers_db
					paper_rec = papers_db.load( args , parsed[ "doi" ] )
				except Exception:
					paper_rec = None
				if paper_rec:
					parsed[ "pdf_path" ] = paper_rec.get( "pdf_path" )
				parsed[ "openalex" ] = _load_openalex_meta( args , parsed[ "doi" ] )
		rows.append( parsed )
		# Hashtag histogram : tokens that look like "#foo" only.
		for tok in ( parsed[ "hashtags" ] or "" ).split():
			if tok.startswith( "#" ) and len( tok ) > 1:
				tag_counter[ tok ] += 1
	return rows , ordered_headers , tag_counter


# ---------------------------------------------------------------------------
# Workbook write
# ---------------------------------------------------------------------------

# Per-column widths ( chars ). Summaries get the most room.
_BASE_COLUMNS = (
	( "DOI"      ,  32 ) ,
	( "Title"    ,  50 ) ,
	( "Hashtags" ,  35 ) ,
)
_SUBSECTION_COL_WIDTH = 60
_FULL_SUMMARY_WIDTH   = 80


def _write_workbook( out_path , rows , ordered_headers , tag_counter ,
                     section_display , extra_columns=None ):
	wb = Workbook()
	ws = wb.active
	ws.title = "Summaries"

	# Build the column list : DOI , Title , Hashtags , < discovered
	# subsections > , < extra config-driven columns > , Full Summary.
	columns = list( _BASE_COLUMNS )
	for h in ordered_headers:
		columns.append( ( h , _SUBSECTION_COL_WIDTH ) )
	extra_names = [ spec[ "name" ] for spec in ( extra_columns or [] ) ]
	for name in extra_names:
		# Extra columns get the same width as subsection columns ; their
		# snippets are sentence-bounded so they wrap well.
		columns.append( ( name , _SUBSECTION_COL_WIDTH ) )
	columns.append( ( "Full Summary" , _FULL_SUMMARY_WIDTH ) )
	header_labels = [ c[ 0 ] for c in columns ]

	# Header row.
	ws.append( header_labels )
	header_font = Font( bold=True , color="FFFFFF" )
	header_fill = PatternFill( "solid" , fgColor="046A38" )
	header_align = Alignment( horizontal="center" , vertical="center" ,
	                          wrap_text=True )
	for cell in ws[ 1 ]:
		cell.font      = header_font
		cell.fill      = header_fill
		cell.alignment = header_align

	# Column widths.
	for i , ( _ , width ) in enumerate( columns , 1 ):
		ws.column_dimensions[ get_column_letter( i ) ].width = width

	# Reusable styles for data rows.
	wrap_top = Alignment( wrap_text=True , vertical="top" )
	link_top = Alignment( wrap_text=False , vertical="top" )

	for row in rows:
		doi      = row[ "doi" ]       or ""
		title    = row[ "title" ]     or ""
		hashtags = row[ "hashtags" ]  or ""
		sub_map  = row.get( "sub_map" , {} )
		full     = row[ "full_summary" ] or ""

		cells = [
			_scrub( doi ) ,
			_scrub( title ) ,
			_scrub( hashtags ) ,
		]
		for h in ordered_headers:
			cells.append( _scrub( sub_map.get( h , "" ) ) )
		extras_map = row.get( "extras" ) or {}
		for name in extra_names:
			cells.append( _scrub( extras_map.get( name , "" ) ) )
		cells.append( _scrub( full ) )

		ws.append( cells )
		r = ws.max_row

		# DOI hyperlink.
		c_doi = ws.cell( row=r , column=1 )
		if doi:
			c_doi.hyperlink = f"https://doi.org/{doi}"
			c_doi.style     = "Hyperlink"
		c_doi.alignment = link_top

		# Wrap every other column.
		for col_idx in range( 2 , len( columns ) + 1 ):
			ws.cell( row=r , column=col_idx ).alignment = wrap_top

	# Freeze header + autofilter for triage-friendliness.
	ws.freeze_panes = "B2"   # also pin DOI column
	last_col = get_column_letter( len( columns ) )
	ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

	# ---- Hashtag histogram sheet ------------------------------------
	ws2 = wb.create_sheet( "Hashtag counts" )
	ws2.append( [ "Hashtag" , "Count" ] )
	for cell in ws2[ 1 ]:
		cell.font      = header_font
		cell.fill      = header_fill
		cell.alignment = header_align
	ws2.column_dimensions[ "A" ].width = 36
	ws2.column_dimensions[ "B" ].width = 10
	for tag , count in tag_counter.most_common():
		ws2.append( [ tag , count ] )
		ws2.cell( row=ws2.max_row , column=2 ).alignment = Alignment(
			horizontal="center" , vertical="top" ,
		)
	ws2.freeze_panes = "A2"
	if ws2.max_row > 1:
		ws2.auto_filter.ref = f"A1:B{ws2.max_row}"

	wb.save( out_path )


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def build_section_xlsx(
	section_dir , section_key , section_display , out_path ,
	args=None , extra_columns=None ,
):
	"""Scan section_dir for *.md , parse , and write the rolled-up xlsx
	at out_path. Returns ( n_rows , n_subsections , n_tags ). Silently
	skips when the section folder has no .md files yet ( early sections
	in a fresh run ). When extra_columns is provided , each spec adds one
	column to the workbook whose cells are sentence-snapped keyword /
	regex snippets pulled from BOTH the LLM summary and the resolved
	section_text ( see config/rollup-extra-columns.yaml )."""
	rows , ordered_headers , tag_counter = _collect_rows(
		section_dir , args=args , section_key=section_key ,
		extra_columns=extra_columns ,
	)
	if not rows:
		print(
			f"ROLLUP    :: {section_key:<13} no .md summaries found under "
			f"{section_dir} ; skipped xlsx"
		)
		return 0 , 0 , 0
	_write_workbook(
		out_path , rows , ordered_headers , tag_counter , section_display ,
		extra_columns=extra_columns ,
	)
	extra_n = len( extra_columns or () )
	print(
		f"ROLLUP    :: {section_key:<13} wrote {len( rows )} rows x "
		f"{len( ordered_headers )} subsection cols + {extra_n} extra cols + "
		f"{len( tag_counter )} hashtag rows -> {out_path}"
	)

	# Per-extra-column markdown list : one .md per column , next to the
	# xlsx , sorted by OpenAlex publication_date desc. Cheap : the
	# per-row openalex / pdf-path lookup already happened during
	# _collect_rows ( only for rows that matched ) , so this is a pure
	# in-memory walk + write.
	if extra_columns:
		for spec in extra_columns:
			slug    = _slug_for_filename( spec[ "name" ] )
			md_path = out_path.parent.joinpath( f"{section_key}-{slug}.md" )
			n_matched = _render_extra_column_md(
				rows , spec , section_key , section_display ,
				md_path , total_n=len( rows ) ,
			)
			if n_matched:
				print(
					f"ROLLUP    :: {section_key:<13} extra-column "
					f"{spec[ 'name' ]!r}: {n_matched} matched papers "
					f"-> {md_path}"
				)
			else:
				print(
					f"ROLLUP    :: {section_key:<13} extra-column "
					f"{spec[ 'name' ]!r}: 0 matched papers ; no .md written"
				)

	return len( rows ) , len( ordered_headers ) , len( tag_counter )


# ---------------------------------------------------------------------------
# CLI entry point ( ` prma rollup [section] ` )
# ---------------------------------------------------------------------------

def _resolve_sections( args ):
	"""Map args.rollup_section -> [ ( key , display ) , ... ]. 'all' is
	every summarizable section ; otherwise a single section by key or
	display name. Same shape as src/tasks/summarize._resolve_sections
	so the two commands always understand the same vocabulary."""
	from ..pdf import section_text as ST
	raw = ( getattr( args , "rollup_section" , None ) or "all" ).lower()
	all_sections = ST.list_sections()
	if raw == "all":
		return all_sections
	wanted = raw.replace( " " , "_" )
	for key , display in all_sections:
		if key == wanted or display.lower().replace( " " , "_" ) == wanted:
			return ( ( key , display ) , )
	valid = ", ".join( k for k , _ in all_sections )
	raise SystemExit(
		f"prma rollup : unknown section {raw!r} ; "
		f"valid choices : all , {valid}"
	)


def run( args ):
	"""Multi-section orchestrator. Walks the resolved section list ,
	calls build_section_xlsx for each section folder that exists , and
	prints a one-line summary at the end. Loads extra-column specs from
	config/rollup-extra-columns.yaml once and reuses them across every
	section."""
	sections      = _resolve_sections( args )
	extra_columns = _load_extra_columns_config( args.config )
	out_root      = args.output.joinpath( "summaries" )
	out_root.mkdir( parents=True , exist_ok=True )

	section_label = ", ".join( k for k , _ in sections )
	extras_label  = (
		f"[{', '.join( c['name'] for c in extra_columns )}]"
		if extra_columns else "(none)"
	)
	print(
		f"ROLLUP    :: sections=[{section_label}] "
		f"extra-columns={extras_label} -> {out_root}"
	)

	n_done , n_empty , n_failed = 0 , 0 , 0
	for key , display in sections:
		section_dir = out_root.joinpath( key )
		out_path    = out_root.joinpath( f"{key}.xlsx" )
		if not section_dir.exists():
			print(
				f"ROLLUP    :: {key:<13} no folder at {section_dir} "
				f"( run ` prma summarize {key} ` first ) ; skipped"
			)
			n_empty += 1
			continue
		try:
			n_rows , _ , _ = build_section_xlsx(
				section_dir , key , display , out_path ,
				args=args , extra_columns=extra_columns ,
			)
		except Exception as e:
			print( f"ROLLUP    :: {key:<13} failed ( {e} )" )
			n_failed += 1
			continue
		if n_rows == 0:
			n_empty += 1
		else:
			n_done += 1

	print(
		f"ROLLUP    :: done ( wrote={n_done} empty={n_empty} failed={n_failed} )"
	)
